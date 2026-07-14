from __future__ import annotations

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from bot.database.models.user import User


_HEADER_FILL = PatternFill("solid", fgColor="1C2128")
_HEADER_FONT = Font(bold=True, color="79C0FF")
_DATE_FMT = "%Y-%m-%d %H:%M"


def _build_workbook(users: list[User], sheet_title: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    headers = [
        "Telegram ID", "Name", "Username",
        "Language", "VIP", "Wallet Balance",
        "Banned", "Joined", "Last Active",
    ]

    # Style headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, user in enumerate(users, start=2):
        row = [
            user.telegram_id,
            user.first_name,
            f"@{user.username}" if user.username else "",
            user.language.value.upper(),
            "Yes" if user.is_vip else "No",
            round(user.wallet_balance, 2),
            "Yes" if user.is_banned else "No",
            user.created_at.strftime(_DATE_FMT) if user.created_at else "",
            user.last_active.strftime(_DATE_FMT) if user.last_active else "",
        ]
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    return wb


def export_users_to_bytes(users: list[User], segment: str = "all") -> bytes:
    """
    Builds an in-memory Excel file and returns its raw bytes.

    segment: "all" | "vip" | "banned"
    """
    title_map = {
        "all": "All Users",
        "vip": "VIP Members",
        "banned": "Banned Users",
    }
    wb = _build_workbook(users, title_map.get(segment, "Export"))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
